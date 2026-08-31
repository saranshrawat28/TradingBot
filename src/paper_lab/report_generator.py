"""
Weekly & Multi-Week Self-Testing Accuracy Report and Signal Diagnostic Generator.
Analyzes win rates, P&L, execution metrics, and pinpoints exact indicator failure modes.
"""

import json
from datetime import datetime, date, timedelta
from pathlib import Path
from typing import Dict, List, Any, Optional
import config

from src.paper_lab.paper_db import PaperDB
from src.paper_lab.lab_config import LabConfig
from src.utils.helpers import get_ist_now, format_currency_inr

REPORTS_DIR = config.STORAGE_DIR / "paper_lab_reports"

class ReportGenerator:
    """Generates comprehensive accuracy audits and self-correcting diagnostic reports."""

    @classmethod
    def generate_report(
        cls,
        days_lookback: int = 7,
        end_date: Optional[str] = None
    ) -> Dict[str, Any]:
        """
        Generates and saves a full accuracy and signal diagnostic report.
        """
        REPORTS_DIR.mkdir(parents=True, exist_ok=True)
        now = get_ist_now()

        end_dt = datetime.strptime(end_date, "%Y-%m-%d") if end_date else now
        start_dt = end_dt - timedelta(days=days_lookback - 1)

        start_str = start_dt.strftime("%Y-%m-%d")
        end_str = end_dt.strftime("%Y-%m-%d")

        records = PaperDB.get_outcomes_for_range(start_str, end_str)

        total_picks = len(records)
        report_title = f"ApexTrade Paper Lab — {'Weekly' if days_lookback <= 7 else 'Multi-Week'} Accuracy Report ({start_str} to {end_str})"

        if total_picks == 0:
            empty_report = {
                "title": report_title,
                "period": {"start": start_str, "end": end_str, "days": days_lookback},
                "total_picks": 0,
                "message": f"No evaluated paper trading outcomes found between {start_str} and {end_str}."
            }
            return empty_report

        # 1. Financial & Win-Rate Aggregates
        total_notional = sum(float(r.get("allocated_capital", LabConfig.DAILY_CAPITAL_PER_PICK)) for r in records)
        total_pnl = sum(float(r.get("pnl_rs", 0.0)) for r in records)
        net_return_pct = round((total_pnl / total_notional) * 100.0, 2) if total_notional > 0 else 0.0

        t1_hits = [r for r in records if r.get("exit_type") == "T1_HIT"]
        t2_hits = [r for r in records if r.get("exit_type") == "T2_HIT"]
        sl_hits = [r for r in records if r.get("exit_type") == "SL_HIT"]
        eod_closes = [r for r in records if r.get("exit_type") == "EOD_CLOSE"]

        winners = [r for r in records if float(r.get("pnl_rs", 0.0)) > 0]
        losers = [r for r in records if float(r.get("pnl_rs", 0.0)) < 0]
        breakeven = [r for r in records if float(r.get("pnl_rs", 0.0)) == 0]

        win_count = len(winners)
        loss_count = len(losers)
        win_rate = round((win_count / total_picks) * 100.0, 1)
        loss_rate = round((loss_count / total_picks) * 100.0, 1)

        total_gains = sum(float(r.get("pnl_rs", 0.0)) for r in winners)
        total_losses = abs(sum(float(r.get("pnl_rs", 0.0)) for r in losers))

        avg_win = round(total_gains / win_count, 2) if win_count > 0 else 0.0
        avg_loss = round(total_losses / loss_count, 2) if loss_count > 0 else 0.0
        profit_factor = round(total_gains / total_losses, 2) if total_losses > 0 else (99.0 if total_gains > 0 else 1.0)

        # Average holding time / bars
        avg_bars_winners = round(sum(int(r.get("bars_held", 0)) for r in winners) / win_count, 1) if win_count > 0 else 0
        avg_bars_losers = round(sum(int(r.get("bars_held", 0)) for r in losers) / loss_count, 1) if loss_count > 0 else 0

        # 2. Signal Failure Diagnostics (Where the Recommender Lagged)
        losing_diagnostics = {
            "total_losers": loss_count,
            "rsi_overbought_entries": 0,
            "weak_volume_entries": 0,
            "late_vwap_entries": 0,
            "low_adx_chop_entries": 0,
            "details": []
        }

        for r in losers:
            sb = r.get("score_breakdown") or {}
            rsi = float(sb.get("rsi", 50.0))
            rvol = float(sb.get("rvol", 1.0))
            vwap_dist = float(sb.get("vwap_sigma_dist", 0.0))
            adx = float(sb.get("adx", 25.0))

            is_rsi_fail = rsi > LabConfig.MAX_ENTRY_RSI_THRESHOLD
            is_rvol_fail = rvol < LabConfig.MIN_ENTRY_RVOL_THRESHOLD
            is_vwap_fail = vwap_dist > LabConfig.MAX_VWAP_SIGMA_THRESHOLD
            is_adx_fail = adx < LabConfig.MIN_ADX_TREND_THRESHOLD

            if is_rsi_fail: losing_diagnostics["rsi_overbought_entries"] += 1
            if is_rvol_fail: losing_diagnostics["weak_volume_entries"] += 1
            if is_vwap_fail: losing_diagnostics["late_vwap_entries"] += 1
            if is_adx_fail: losing_diagnostics["low_adx_chop_entries"] += 1

            losing_diagnostics["details"].append({
                "symbol": r["symbol"],
                "date": r["pick_date"],
                "pnl_rs": r["pnl_rs"],
                "pnl_pct": r["pnl_pct"],
                "exit_type": r["exit_type"],
                "entry_rsi": rsi,
                "entry_rvol": rvol,
                "entry_vwap_sigma": vwap_dist,
                "entry_adx": adx,
                "reasons_flagged": [
                    m for m, f in [
                        (f"RSI Overbought ({rsi:.1f} > {LabConfig.MAX_ENTRY_RSI_THRESHOLD})", is_rsi_fail),
                        (f"Weak Relative Volume ({rvol:.2f} < {LabConfig.MIN_ENTRY_RVOL_THRESHOLD})", is_rvol_fail),
                        (f"Late Entry Above VWAP ({vwap_dist:+.2f}σ > {LabConfig.MAX_VWAP_SIGMA_THRESHOLD}σ)", is_vwap_fail),
                        (f"Low ADX Chop Regime ({adx:.1f} < {LabConfig.MIN_ADX_TREND_THRESHOLD})", is_adx_fail)
                    ] if f
                ]
            })

        # Calculate percentages
        rsi_fail_pct = round((losing_diagnostics["rsi_overbought_entries"] / loss_count) * 100.0, 1) if loss_count > 0 else 0.0
        rvol_fail_pct = round((losing_diagnostics["weak_volume_entries"] / loss_count) * 100.0, 1) if loss_count > 0 else 0.0
        vwap_fail_pct = round((losing_diagnostics["late_vwap_entries"] / loss_count) * 100.0, 1) if loss_count > 0 else 0.0
        adx_fail_pct = round((losing_diagnostics["low_adx_chop_entries"] / loss_count) * 100.0, 1) if loss_count > 0 else 0.0

        # 3. Actionable Optimization Recommendations
        recommendations = []
        if loss_count > 0:
            if rsi_fail_pct >= 40.0:
                recommendations.append(
                    f"🎯 **RSI Filter Calibration**: {losing_diagnostics['rsi_overbought_entries']}/{loss_count} ({rsi_fail_pct}%) of losing trades entered when RSI was above {LabConfig.MAX_ENTRY_RSI_THRESHOLD}. Consider tightening the maximum allowable RSI at entry in StockAdvisor to ≤ 62.0."
                )
            if rvol_fail_pct >= 40.0:
                recommendations.append(
                    f"📊 **Volume Confirmation**: {losing_diagnostics['weak_volume_entries']}/{loss_count} ({rvol_fail_pct}%) of losing trades lacked institutional volume (RVol < {LabConfig.MIN_ENTRY_RVOL_THRESHOLD:.2f}). Require RVol ≥ 1.10 before triggering momentum BUY verdicts."
                )
            if vwap_fail_pct >= 30.0:
                recommendations.append(
                    f"📍 **VWAP Overextension Protection**: {losing_diagnostics['late_vwap_entries']}/{loss_count} ({vwap_fail_pct}%) of losing trades were entered more than {LabConfig.MAX_VWAP_SIGMA_THRESHOLD}σ extended from intraday VWAP. Implement a strict 'No-Chase' rule when price exceeds VWAP + 0.35σ."
                )
            if adx_fail_pct >= 40.0:
                recommendations.append(
                    f"⚡ **Regime Filter**: {losing_diagnostics['low_adx_chop_entries']}/{loss_count} ({adx_fail_pct}%) of losing trades occurred in low ADX (< {LabConfig.MIN_ADX_TREND_THRESHOLD:.0f}) non-trending environments. Gate momentum breakouts behind ADX ≥ 22.0."
                )
            if not recommendations:
                recommendations.append("✨ Strategy parameters performed well within baseline tolerance. Continue monitoring sample size.")
        else:
            recommendations.append("🌟 100% win rate during this period! No negative signal lags detected.")

        # 4. Low Sample Size Warning
        is_low_sample = (total_picks < 15 or loss_count < 8)
        sample_warning = None
        if is_low_sample:
            sample_warning = f"⚠️ PRELIMINARY DATA: Total sample size is small ({total_picks} total trades, {loss_count} losses). Treat diagnostic percentages as directional trends rather than definitive statistical proof. Continue paper testing for 2–4 full weeks."

        # Filter out any synthetic test artifacts
        real_records = [r for r in records if not r.get("symbol", "").startswith(("TEST_", "DIAG_"))]
        if not real_records:
            real_records = records

        # 5. Top Winners & Worst Losers (Strictly positive for winners, strictly negative for losers)
        sorted_by_pnl = sorted(real_records, key=lambda x: float(x.get("pnl_rs", 0.0)), reverse=True)
        top_winners = [r for r in sorted_by_pnl if float(r.get("pnl_rs", 0.0)) > 0][:5]
        worst_losers = [r for r in sorted(real_records, key=lambda x: float(x.get("pnl_rs", 0.0))) if float(r.get("pnl_rs", 0.0)) < 0][:5]

        # 6. Assemble Full Structured Report
        report_data = {
            "title": report_title,
            "period": {
                "start": start_str,
                "end": end_str,
                "days": days_lookback
            },
            "config_version": LabConfig.CONFIG_VERSION,
            "sample_warning": sample_warning,
            "financial_summary": {
                "total_notional_deployed_rs": total_notional,
                "net_realized_pnl_rs": total_pnl,
                "net_return_pct": net_return_pct,
                "profit_factor": profit_factor,
                "total_gains_rs": total_gains,
                "total_losses_rs": total_losses,
                "avg_winner_rs": avg_win,
                "avg_loser_rs": avg_loss
            },
            "prediction_accuracy": {
                "total_picks": total_picks,
                "winning_picks": win_count,
                "losing_picks": loss_count,
                "breakeven_picks": len(breakeven),
                "win_rate_pct": win_rate,
                "loss_rate_pct": loss_rate,
                "t1_hit_count": len(t1_hits),
                "t2_hit_count": len(t2_hits),
                "sl_hit_count": len(sl_hits),
                "eod_close_count": len(eod_closes),
                "avg_bars_to_win": avg_bars_winners,
                "avg_bars_to_loss": avg_bars_losers
            },
            "signal_diagnostics": {
                "total_failures_analyzed": loss_count,
                "rsi_fail_count": losing_diagnostics["rsi_overbought_entries"],
                "rsi_fail_pct": rsi_fail_pct,
                "rvol_fail_count": losing_diagnostics["weak_volume_entries"],
                "rvol_fail_pct": rvol_fail_pct,
                "vwap_fail_count": losing_diagnostics["late_vwap_entries"],
                "vwap_fail_pct": vwap_fail_pct,
                "adx_fail_count": losing_diagnostics["low_adx_chop_entries"],
                "adx_fail_pct": adx_fail_pct,
                "failing_trade_breakdown": losing_diagnostics["details"]
            },
            "recommendations": recommendations,
            "top_winners": top_winners,
            "worst_losers": worst_losers,
            "records": real_records
        }

        # 7. Generate Clean Markdown Text
        md_lines = [
            f"# 🧪 {report_title}",
            f"**Evaluation Period**: `{start_str}` to `{end_str}` | **Active Engine Version**: `{LabConfig.CONFIG_VERSION}`\n",
        ]

        if sample_warning:
            md_lines.append(f"> {sample_warning}\n")

        md_lines.extend([
            "## 📊 Financial & Performance Summary",
            f"• **Total Notional Capital Deployed**: `{format_currency_inr(total_notional)}` (₹20,000 flat per trade)",
            f"• **Net Realized P&L**: **{'+' if total_pnl > 0 else ''}{format_currency_inr(total_pnl)}** (`{net_return_pct:+.2f}%` on capital)",
            f"• **Profit Factor**: `{profit_factor:.2f}` (Total Gains: `{format_currency_inr(total_gains)}` | Total Losses: `{format_currency_inr(total_losses)}`)",
            f"• **Avg Win / Avg Loss**: `{format_currency_inr(avg_win)}` / `-{format_currency_inr(avg_loss)}`\n",
            "## 🎯 Recommendation Accuracy Breakdown",
            f"• **Total Stock Recommendations**: `{total_picks}`",
            f"• ✅ **Winning Trades (Hit Target 1 / 2)**: `{win_count}` (`{win_rate}%`)",
            f"• ❌ **Stopped Out (Hit Stop-Loss)**: `{loss_count}` (`{loss_rate}%`)",
            f"• ↩️ **EOD Timeouts (Closed at 3:25 PM)**: `{len(eod_closes)}` (`{round(len(eod_closes)/total_picks*100, 1) if total_picks else 0.0}%`)",
            f"• 🎯 **Target 1 Hits**: `{len(t1_hits)}` | 🚀 **Target 2 Hits**: `{len(t2_hits)}` | 🛑 **SL Hits**: `{len(sl_hits)}`\n",
            "## 🔬 Signal Lag Diagnostics (Where the Recommender Failed)",
            f"*Analyzed all `{loss_count}` losing trade setups to identify recurring failure patterns:*\n",
            f"1. **RSI Overbought Exhaustion**: `{losing_diagnostics['rsi_overbought_entries']}/{loss_count}` losing trades (`{rsi_fail_pct}%`) were entered with RSI > {LabConfig.MAX_ENTRY_RSI_THRESHOLD:.0f}.",
            f"2. **Weak Volume Confirmation**: `{losing_diagnostics['weak_volume_entries']}/{loss_count}` losing trades (`{rvol_fail_pct}%`) had low relative volume (RVol < {LabConfig.MIN_ENTRY_RVOL_THRESHOLD:.2f}).",
            f"3. **Late VWAP Chasing**: `{losing_diagnostics['late_vwap_entries']}/{loss_count}` losing trades (`{vwap_fail_pct}%`) entered > {LabConfig.MAX_VWAP_SIGMA_THRESHOLD:.2f}σ above VWAP.",
            f"4. **Low ADX Chop Breakouts**: `{losing_diagnostics['low_adx_chop_entries']}/{loss_count}` losing trades (`{adx_fail_pct}%`) occurred in non-trending regimes (ADX < {LabConfig.MIN_ADX_TREND_THRESHOLD:.0f}).\n",
            "## 💡 Actionable Improvement Recommendations"
        ])

        for rec in recommendations:
            md_lines.append(f"• {rec}")

        md_lines.append("\n## 🏆 Best & Worst Recommendations")
        md_lines.append("### 🌟 Top Winners")
        if top_winners:
            for i, w in enumerate(top_winners, 1):
                pnl_val = float(w.get('pnl_rs', 0.0))
                pnl_pct_val = float(w.get('pnl_pct', 0.0))
                md_lines.append(f"{i}. **{w['symbol']}** ({w['pick_date']}): `{w['exit_type']}` @ `₹{w['exit_price']:,.2f}` | P&L: **+{format_currency_inr(pnl_val)}** (`+{pnl_pct_val:.2f}%`)")
        else:
            md_lines.append("• *No winning target hits recorded in this period (trades either hit SL or timed out at 3:25 PM EOD).*")

        md_lines.append("\n### 📉 Worst Losers")
        if worst_losers:
            for i, l in enumerate(worst_losers, 1):
                pnl_val = float(l.get('pnl_rs', 0.0))
                pnl_pct_val = float(l.get('pnl_pct', 0.0))
                md_lines.append(f"{i}. **{l['symbol']}** ({l['pick_date']}): `{l['exit_type']}` @ `₹{l['exit_price']:,.2f}` | P&L: **{format_currency_inr(pnl_val)}** (`{pnl_pct_val:.2f}%`)")
        else:
            md_lines.append("• *No losing trades recorded in this period (100% win rate).*")

        markdown_text = "\n".join(md_lines)
        report_data["markdown_text"] = markdown_text

        # 8. Save to storage (TXT, JSON, XLSX, CSV)
        report_file_txt = REPORTS_DIR / f"report_{end_str}_{days_lookback}d.txt"
        report_file_json = REPORTS_DIR / f"report_{end_str}_{days_lookback}d.json"
        report_file_xlsx = REPORTS_DIR / f"report_{end_str}_{days_lookback}d.xlsx"
        report_file_csv = REPORTS_DIR / f"report_{end_str}_{days_lookback}d.csv"

        with open(report_file_txt, "w", encoding="utf-8") as f:
            f.write(markdown_text)

        with open(report_file_json, "w", encoding="utf-8") as f:
            json.dump(report_data, f, indent=2, default=str)

        try:
            xlsx_bytes = cls.export_report_to_excel_bytes(report_data)
            with open(report_file_xlsx, "wb") as f:
                f.write(xlsx_bytes)
        except Exception as e:
            pass

        try:
            csv_text = cls.export_report_to_csv_string(report_data)
            with open(report_file_csv, "w", encoding="utf-8") as f:
                f.write(csv_text)
        except Exception as e:
            pass

        try:
            print(f"[ReportGenerator] Report successfully generated and saved to {report_file_txt}")
        except Exception:
            pass
        return report_data

    @classmethod
    def export_report_to_excel_bytes(cls, report_data: Dict[str, Any]) -> bytes:
        """
        Exports the report data into a professionally styled, multi-tab Microsoft Excel (.xlsx) workbook.
        """
        import io
        import openpyxl
        from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
        from openpyxl.utils import get_column_letter

        wb = openpyxl.Workbook()

        header_fill = PatternFill(start_color="1E3A8A", end_color="1E3A8A", fill_type="solid")
        header_font = Font(name="Calibri", size=11, bold=True, color="FFFFFF")
        title_font = Font(name="Calibri", size=13, bold=True, color="1E3A8A")
        bold_font = Font(name="Calibri", size=11, bold=True)
        regular_font = Font(name="Calibri", size=11)

        win_fill = PatternFill(start_color="DCFCE7", end_color="DCFCE7", fill_type="solid")
        loss_fill = PatternFill(start_color="FEE2E2", end_color="FEE2E2", fill_type="solid")
        neutral_fill = PatternFill(start_color="FEF3C7", end_color="FEF3C7", fill_type="solid")

        thin_border = Border(
            left=Side(style='thin', color='CBD5E1'),
            right=Side(style='thin', color='CBD5E1'),
            top=Side(style='thin', color='CBD5E1'),
            bottom=Side(style='thin', color='CBD5E1')
        )
        center_align = Alignment(horizontal='center', vertical='center')
        left_align = Alignment(horizontal='left', vertical='center')
        right_align = Alignment(horizontal='right', vertical='center')

        # -------------------------------------------------------------
        # SHEET 1: Session_Trade_Details
        # -------------------------------------------------------------
        ws1 = wb.active
        ws1.title = "Session_Trade_Details"
        ws1.views.sheetView[0].showGridLines = True

        period_str = f"{report_data.get('period', {}).get('start', '')} to {report_data.get('period', {}).get('end', '')}"
        ws1["A1"] = f"APEXTRADE ACCURACY LAB - TRADE EXECUTION DETAILS ({period_str})"
        ws1["A1"].font = title_font
        ws1.merge_cells("A1:U1")
        ws1.row_dimensions[1].height = 26

        headers = [
            "Date", "Symbol", "Company Name", "Quant Score", "Setup Grade",
            "Signal Time", "Signal Price (₹)", "Entry Time", "Entry Price (₹)",
            "Target 1 (₹)", "Target 2 (₹)", "Stop-Loss (₹)", "Allocated Capital (₹)",
            "Quantity", "Outcome Status", "Exit Time", "Exit Price (₹)",
            "Realized P&L (₹)", "Return (%)", "Bars Held (Mins)", "Key Technical Catalysts"
        ]

        ws1.append([])
        ws1.append(headers)
        ws1.row_dimensions[3].height = 26

        for col_idx, _ in enumerate(headers, 1):
            cell = ws1.cell(row=3, column=col_idx)
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = center_align
            cell.border = thin_border

        records = report_data.get("records", [])
        for row_idx, r in enumerate(records, 4):
            pnl = float(r.get("pnl_rs", 0.0))
            pnl_pct = float(r.get("pnl_pct", 0.0))
            ext_type = str(r.get("exit_type", "ACTIVE"))
            top_sigs = r.get("top_signals", [])
            sig_str = ", ".join(top_sigs[:2]) if isinstance(top_sigs, list) else str(top_sigs or "")

            row_data = [
                r.get("pick_date", ""),
                r.get("symbol", ""),
                r.get("display_name", r.get("symbol", "")),
                f"{float(r.get('advisor_score', 0)):.1f}/10",
                r.get("setup_grade", "GRADE A"),
                r.get("signal_time", "08:50:00"),
                float(r.get("signal_price", 0.0)),
                r.get("entry_time", "09:15:00"),
                float(r.get("entry_price", 0.0)),
                float(r.get("target_1", 0.0)),
                float(r.get("target_2", 0.0)),
                float(r.get("stop_loss", 0.0)),
                float(r.get("allocated_capital", 20000.0)),
                int(r.get("quantity", 1)),
                ext_type,
                r.get("exit_time", "15:15:00"),
                float(r.get("exit_price", 0.0)),
                pnl,
                pnl_pct / 100.0,
                int(r.get("bars_held", 0)),
                sig_str
            ]
            ws1.append(row_data)
            ws1.row_dimensions[row_idx].height = 22

            for col_idx in range(1, len(headers) + 1):
                c = ws1.cell(row=row_idx, column=col_idx)
                c.font = regular_font
                c.border = thin_border

                if col_idx in [7, 9, 10, 11, 12, 13, 17, 18]:
                    c.number_format = '₹#,##0.00'
                    c.alignment = right_align
                elif col_idx == 14:
                    c.number_format = '#,##0'
                    c.alignment = center_align
                elif col_idx == 19:
                    c.number_format = '+0.00%;-0.00%;0.00%'
                    c.alignment = right_align
                elif col_idx in [1, 6, 8, 15, 16, 20]:
                    c.alignment = center_align
                else:
                    c.alignment = left_align

                if "HIT" in ext_type and ("T1" in ext_type or "T2" in ext_type):
                    if col_idx in [15, 18, 19]:
                        c.fill = win_fill
                        c.font = bold_font
                elif "SL" in ext_type or pnl < 0:
                    if col_idx in [15, 18, 19]:
                        c.fill = loss_fill
                        c.font = bold_font
                elif "EOD" in ext_type:
                    if col_idx in [15, 18, 19]:
                        c.fill = neutral_fill

        for col in ws1.columns:
            max_len = max(len(str(cell.value or '')) for cell in col)
            col_letter = get_column_letter(col[0].column)
            ws1.column_dimensions[col_letter].width = max(max_len + 4, 12)

        # -------------------------------------------------------------
        # SHEET 2: Executive_Summary
        # -------------------------------------------------------------
        ws2 = wb.create_sheet(title="Executive_Summary")
        ws2.views.sheetView[0].showGridLines = True

        ws2["A1"] = "APEXTRADE PAPER LAB - EXECUTIVE PERFORMANCE AUDIT"
        ws2["A1"].font = title_font
        ws2.merge_cells("A1:C1")
        ws2.row_dimensions[1].height = 26

        fin = report_data.get("financial_summary", {})
        acc = report_data.get("prediction_accuracy", {})
        period = report_data.get("period", {})

        summary_items = [
            ("Evaluation Start Date", period.get("start", ""), "@"),
            ("Evaluation End Date", period.get("end", ""), "@"),
            ("Active Algorithm Config", report_data.get("config_version", "v1.0.0"), "@"),
            ("Total Stock Recommendations", acc.get("total_picks", 0), "#,##0"),
            ("Winning Recommendations (T1/T2)", acc.get("winning_picks", 0), "#,##0"),
            ("Stopped Out Recommendations (SL)", acc.get("losing_picks", 0), "#,##0"),
            ("EOD Timeouts (Closed 3:25 PM)", acc.get("eod_close_count", 0), "#,##0"),
            ("Recommendation Win Rate (%)", (acc.get("win_rate_pct", 0.0) / 100.0), "0.0%"),
            ("Total Capital Deployed (₹)", fin.get("total_notional_deployed_rs", 0.0), "₹#,##0.00"),
            ("Gross Realized Profit (₹)", fin.get("total_gains_rs", 0.0), "₹#,##0.00"),
            ("Gross Realized Loss (₹)", fin.get("total_losses_rs", 0.0), "₹#,##0.00"),
            ("Net Realized P&L (₹)", fin.get("net_realized_pnl_rs", 0.0), "₹#,##0.00"),
            ("Net Return on Capital (%)", (fin.get("net_return_pct", 0.0) / 100.0), "+0.00%;-0.00%;0.00%"),
            ("Profit Factor", fin.get("profit_factor", 0.0), "0.00"),
            ("Average Winning Trade (₹)", fin.get("avg_winner_rs", 0.0), "₹#,##0.00"),
            ("Average Losing Trade (₹)", fin.get("avg_loser_rs", 0.0), "₹#,##0.00"),
            ("Target 1 Hits Count", acc.get("t1_hit_count", 0), "#,##0"),
            ("Target 2 Hits Count", acc.get("t2_hit_count", 0), "#,##0"),
            ("Stop-Loss Hits Count", acc.get("sl_hit_count", 0), "#,##0")
        ]

        ws2.append([])
        ws2.append(["Metric / Performance KPI", "Value", "Notes / Context"])
        ws2.row_dimensions[3].height = 24
        for c_idx in range(1, 4):
            cell = ws2.cell(row=3, column=c_idx)
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = center_align
            cell.border = thin_border

        for idx, (label, val, fmt) in enumerate(summary_items, 4):
            ws2.append([label, val, ""])
            ws2.row_dimensions[idx].height = 20
            c1 = ws2.cell(row=idx, column=1)
            c2 = ws2.cell(row=idx, column=2)
            c3 = ws2.cell(row=idx, column=3)

            c1.font = bold_font
            c1.border = thin_border
            c2.font = regular_font
            c2.number_format = fmt
            c2.alignment = right_align
            c2.border = thin_border
            c3.border = thin_border

            if "Net Realized P&L" in label:
                c2.font = bold_font
                c2.fill = win_fill if float(fin.get("net_realized_pnl_rs", 0)) >= 0 else loss_fill

        ws2.column_dimensions["A"].width = 38
        ws2.column_dimensions["B"].width = 22
        ws2.column_dimensions["C"].width = 25

        # -------------------------------------------------------------
        # SHEET 3: Signal_Diagnostics
        # -------------------------------------------------------------
        ws3 = wb.create_sheet(title="Signal_Diagnostics")
        ws3.views.sheetView[0].showGridLines = True
        ws3["A1"] = "SIGNAL LAG DIAGNOSTICS & SYSTEM RECOMMENDATIONS"
        ws3["A1"].font = title_font
        ws3.merge_cells("A1:C1")
        ws3.row_dimensions[1].height = 26

        diag = report_data.get("signal_diagnostics", {})
        recs = report_data.get("recommendations", [])

        ws3.append([])
        ws3.append(["Diagnostic Failure Metric", "Count / Ratio", "Failure Rate (%)"])
        for c_idx in range(1, 4):
            cell = ws3.cell(row=3, column=c_idx)
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = center_align
            cell.border = thin_border

        diag_rows = [
            ("Total Losses Analyzed", diag.get("total_failures_analyzed", 0), "100.0%"),
            ("RSI Overbought Exhaustion (> 65)", diag.get("rsi_fail_count", 0), f"{diag.get('rsi_fail_pct', 0.0):.1f}%"),
            ("Weak Relative Volume Flow (< 1.00x)", diag.get("rvol_fail_count", 0), f"{diag.get('rvol_fail_pct', 0.0):.1f}%"),
            ("Late Entry Above VWAP (> 0.40σ)", diag.get("vwap_fail_count", 0), f"{diag.get('vwap_fail_pct', 0.0):.1f}%"),
            ("Low ADX Chop Regime (< 20)", diag.get("adx_fail_count", 0), f"{diag.get('adx_fail_pct', 0.0):.1f}%"),
        ]

        for idx, (lbl, cnt, pct) in enumerate(diag_rows, 4):
            ws3.append([lbl, cnt, pct])
            for c_idx in range(1, 4):
                c = ws3.cell(row=idx, column=c_idx)
                c.font = regular_font
                c.border = thin_border
                if c_idx > 1: c.alignment = center_align

        rec_start = len(diag_rows) + 6
        ws3.cell(row=rec_start, column=1, value="ACTIONABLE OPTIMIZATION RECOMMENDATIONS").font = bold_font
        for r_idx, rec_text in enumerate(recs, rec_start + 1):
            ws3.cell(row=r_idx, column=1, value=f"• {rec_text}").font = regular_font

        ws3.column_dimensions["A"].width = 50
        ws3.column_dimensions["B"].width = 18
        ws3.column_dimensions["C"].width = 20

        buf = io.BytesIO()
        wb.save(buf)
        return buf.getvalue()

    @classmethod
    def export_report_to_csv_string(cls, report_data: Dict[str, Any]) -> str:
        """
        Exports the trade execution records into a standard CSV string.
        """
        import io, csv
        buf = io.StringIO()
        writer = csv.writer(buf)

        headers = [
            "Date", "Symbol", "Company Name", "Quant Score", "Setup Grade",
            "Signal Time", "Signal Price", "Entry Time", "Entry Price",
            "Target 1", "Target 2", "Stop Loss", "Allocated Capital",
            "Quantity", "Outcome Status", "Exit Time", "Exit Price",
            "Realized PnL Rs", "Return Pct", "Bars Held", "Key Catalysts"
        ]
        writer.writerow(headers)

        for r in report_data.get("records", []):
            top_sigs = r.get("top_signals", [])
            sig_str = ", ".join(top_sigs[:2]) if isinstance(top_sigs, list) else str(top_sigs or "")
            writer.writerow([
                r.get("pick_date", ""),
                r.get("symbol", ""),
                r.get("display_name", r.get("symbol", "")),
                float(r.get("advisor_score", 0)),
                r.get("setup_grade", "GRADE A"),
                r.get("signal_time", "08:50:00"),
                float(r.get("signal_price", 0.0)),
                r.get("entry_time", "09:15:00"),
                float(r.get("entry_price", 0.0)),
                float(r.get("target_1", 0.0)),
                float(r.get("target_2", 0.0)),
                float(r.get("stop_loss", 0.0)),
                float(r.get("allocated_capital", 20000.0)),
                int(r.get("quantity", 1)),
                r.get("exit_type", "ACTIVE"),
                r.get("exit_time", "15:15:00"),
                float(r.get("exit_price", 0.0)),
                float(r.get("pnl_rs", 0.0)),
                float(r.get("pnl_pct", 0.0)),
                int(r.get("bars_held", 0)),
                sig_str
            ])
        return buf.getvalue()

    @classmethod
    def list_saved_reports(cls) -> List[Dict[str, Any]]:
        """Lists all historical saved reports in storage."""
        if not REPORTS_DIR.exists():
            return []

        reports = []
        for p in sorted(REPORTS_DIR.glob("*.json"), reverse=True):
            try:
                with open(p, "r", encoding="utf-8") as f:
                    data = json.load(f)
                    reports.append({
                        "file_path": str(p),
                        "file_name": p.name,
                        "title": data.get("title", p.stem),
                        "period": data.get("period", {}),
                        "win_rate": data.get("prediction_accuracy", {}).get("win_rate_pct", 0.0),
                        "net_pnl": data.get("financial_summary", {}).get("net_realized_pnl_rs", 0.0)
                    })
            except Exception:
                pass
        return reports
